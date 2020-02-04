import openpyxl
import requests
from bs4 import BeautifulSoup
import numpy as np
import threading
import time

def getHTML(url: str, headers: dict):
    try:
        response = requests.get(url, headers = headers)
        if response.status_code != 200: 
            print('Error:',response.status_code)
            print(url)
        return response.text
    except Exception as e:
        raise e

def getWorkBook(urlFile: str)->openpyxl.workbook.workbook.Workbook:
    """Load work book .xlsx"""
    try:
        workBook = openpyxl.load_workbook(filename=urlFile)
    except openpyxl.utils.exceptions.InvalidFileException:
        raise SystemExit("Formant file .xls not support. Use formal .xlsx")
    except FileNotFoundError:
        raise SystemExit('File not fount in derectory')
    return workBook

def getWorkSheet(workBook: openpyxl.workbook.workbook.Workbook)->openpyxl.worksheet.worksheet.Worksheet:
    """Load first sheet from book"""
    try:
        sheet = workBook[workBook.sheetnames[0]]
        return sheet
    except TypeError as e:
        sheet = workBook.sheet_by_index(0)
        return e

def appendToFile(urlFile: str, data: dict)->dict:
    """Writing new lines ti the end of the file"""
    workBook = getWorkBook(urlFile)
    sheet = getWorkSheet(workBook)
    numRow = sheet.max_row+1
    try:
        sheet.cell(row=numRow, column=1).value = data['name']
        sheet.cell(row=numRow, column=2).value = data['rating']
        sheet.cell(row=numRow, column=3).value = data['description']
        sheet.cell(row=numRow, column=4).value = data['addres']
        sheet.cell(row=numRow, column=5).value = data['phone']
        sheet.cell(row=numRow, column=6).value = data['website']
        sheet.cell(row=numRow, column=7).value = data['email']
    except Exception as e:
        print(e)
    try:
        workBook.save(urlFile)
        workBook.close()
    except Exception as e:
        print(e)
    return {
        'urlFile':urlFile,
        'All rows:':sheet.max_row,
    }

def parsingPageOrganization(urlOrganization:str,file):
    print(urlOrganization)
    headers={
        "Host": "2gis.ru",
        "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        'Accept-Encoding': "gzip, deflate",
        "Connection": "keep-alive",
        "Cookie": "captcha=eyJhbGciOiJFUzI1NiIsInR5cCI6IkpXVCJ9.eyJpcCI6IjgzLjIyMS4yMDUuMjAxIiwiYXVkIjoicGFyc2VyLWRldGVjdG9yIiwiZXhwIjoxNTgwNjQ2NTQyLCJpc3MiOiJjYXB0Y2hhIn0.toP0OcfESstpYNnhB6zCrVAXARRe0vJFM0csYWSdjSuxqW3Zc8orKegW0UsBTt0eU-hPwamfWk7quAtcBIg9fQ;",
        "Upgrade-Insecure-Requests": "1",
        "Pragma": "no-cache",
        "Cache-Control": "no-cache",
        "TE": "Trailers",
    }
    response = getHTML(urlOrganization,headers)
    soup = BeautifulSoup(response,features="html.parser")
    organizationData = {}
    try:
        organizationData['name'] = soup.find('span',{'class':'_oqoid'}).text
    except Exception as e:
        organizationData['name'] = "non-data"
    try:
        organizationData['rating'] = soup.find('span',{'class':'_1n8h0vx'}).text
    except Exception as e:
        organizationData['rating'] = "non-data"
    try:
        organizationData['description'] = soup.find('pre',{'class':'_1tasqnk'}).text
    except Exception as e:
        organizationData['description'] = "non-data"
    try:
        organizationData['addres'] = soup.find('span',{'class':'_14quei'}).find('a',{'class':'_84s065h'}).text + " " + soup.find('div',{'class':'_1p8iqzw'}).text
    except Exception as e:
        organizationData['addres'] = "non-data"
    try:
        organizationData['phone'] = soup.find('div',{'class':'_b0ke8'}).find('a',{'class':'_84s065h'}).get('href')
    except Exception as e:
        organizationData['phone'] = "non-data"
    try:
        organizationData['website'] = soup.find('div',{'class':'_599hh'}).find('div',{'class':'_49kxlr'}).find('a',{'class':'_13ptbeu'}).get('href')
    except Exception as e:
        organizationData['website'] = "non-data"
    try:
        str = soup.find('div',{'class':'_599hh'}).find_all('a',{'class':'_84s065h'})
        for item in str:
            if 'mailto' in item.get('href'):
                organizationData['email'] = item.text
    except Exception as e:
        organizationData['email'] = "non-data"
    
    print(appendToFile(file,organizationData))
    for item in organizationData.keys():
        print(item,': ',organizationData[item])
    time.sleep(3)

def getOrganizations(urlFile:str):
    workBook = getWorkBook(urlFile)
    sheet = getWorkSheet(workBook)
    rows = []
    for i in range(1,sheet.max_row):
        rows.append(sheet.cell(row=i+1, column=1).value)
    workBook.close()
    return rows

def onetherd(listUrls:list,file):
    for item in listUrls:
        parsingPageOrganization("https://2gis.ru"+item.strip(),file)
        


def main():
    listOrganizations = getOrganizations('data2.xlsx')
    part1, part2, part3 = np.array_split(listOrganizations,3)

    t1 = threading.Thread(target=onetherd,args=(part1,'outdata1.xlsx')) 
    t2 = threading.Thread(target=onetherd,args=(part2,'outdata2.xlsx')) 
    t3 = threading.Thread(target=onetherd,args=(part3,'outdata3.xlsx')) 

    t1.start()
    t2.start()
    t3.start()

    t1.join()
    t2.join()
    t3.join()



if __name__ == '__main__':
    main()
