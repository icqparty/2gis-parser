import requests
from bs4 import BeautifulSoup
import json
import threading
import random
import time
import openpyxl

def getHTML(url: str, headers: dict):
    try:
        response = requests.get(url, headers = headers)
        print(url)
        if response.status_code != 200: 
            print('Error:',response.status_code)
            print(url)
        return response.text
    except Exception as e:
        raise e

def parsingOnePlace(sity:str,coordinates:str):
    organizations = []
    countOrganizations = getCountOrganizations(sity,coordinates)
    if countOrganizations != None:
        countPages = int(countOrganizations.text)//12
        for page in range(1,countPages+1):
            organizations.extend(searchOrganizations(page,sity,coordinates))
        print(appendToFile('data.xlsx',organizations))

def getCountOrganizations(sity:str,coordinates:str):
    url = "https://2gis.ru/{0}/search/страхование?{1}".format(sity,coordinates)
    headers={
        "Host": "2gis.ru",
        "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        'Accept-Encoding': "gzip, deflate",
        "Connection": "keep-alive",
        "Cookie": "captcha=eyJhbGciOiJFUzI1NiIsInR5cCI6IkpXVCJ9.eyJpcCI6IjgzLjIyMS4yMDUuMjAxIiwiYXVkIjoicGFyc2VyLWRldGVjdG9yIiwiZXhwIjoxNTgwNjQ2NTQyLCJpc3MiOiJjYXB0Y2hhIn0.toP0OcfESstpYNnhB6zCrVAXARRe0vJFM0csYWSdjSuxqW3Zc8orKegW0UsBTt0eU-hPwamfWk7quAtcBIg9fQ;",
        "Upgrade-Insecure-Requests": "1",
        "Pragma": "no-cache",
        "Cache-Control": "no-cache",
        "TE": "Trailers",
    }
    response = getHTML(url,headers)
    soup = BeautifulSoup(response,features="html.parser")
    return soup.find('span',{'class':'_1cho7kd9'})

def sarchBranches(urlToBranche:str) -> list:
    headers={
        "Host": "2gis.ru",
        "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        'Accept-Encoding': "gzip, deflate",
        "Connection": "keep-alive",
        "Cookie": "captcha=eyJhbGciOiJFUzI1NiIsInR5cCI6IkpXVCJ9.eyJpcCI6IjgzLjIyMS4yMDUuMjAxIiwiYXVkIjoicGFyc2VyLWRldGVjdG9yIiwiZXhwIjoxNTgwNjQ2NTQyLCJpc3MiOiJjYXB0Y2hhIn0.toP0OcfESstpYNnhB6zCrVAXARRe0vJFM0csYWSdjSuxqW3Zc8orKegW0UsBTt0eU-hPwamfWk7quAtcBIg9fQ;",
        "Upgrade-Insecure-Requests": "1",
        "Pragma": "no-cache",
        "Cache-Control": "no-cache",
        "TE": "Trailers",
    }
    response = getHTML(urlToBranche,headers)
    soup = BeautifulSoup(response)
    branchesHTML = soup.find_all('div', {'class':'_46pzdl'})
    linksToBranches = []
    for item in branchesHTML:
        linksToBranches.append(item.find('a', {'class':'_13ptbeu'}).get('href'))
    return linksToBranches

def searchOrganizations(numberPage,sity:str,coordinates:str):
    if numberPage>1:
        url = "https://2gis.ru/{0}/search/страхование/page/{1}?{2}".format(sity,numberPage,coordinates)
    else:
        url = "https://2gis.ru/{0}/search/страхование?{0}".format(sity,coordinates)
    headers={
        "Host": "2gis.ru",
        "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:72.0) Gecko/20100101 Firefox/72.0",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        'Accept-Encoding': "gzip, deflate",
        "Connection": "keep-alive",
        "Cookie": "captcha=eyJhbGciOiJFUzI1NiIsInR5cCI6IkpXVCJ9.eyJpcCI6IjgzLjIyMS4yMDUuMjAxIiwiYXVkIjoicGFyc2VyLWRldGVjdG9yIiwiZXhwIjoxNTgwNjQ2NTQyLCJpc3MiOiJjYXB0Y2hhIn0.toP0OcfESstpYNnhB6zCrVAXARRe0vJFM0csYWSdjSuxqW3Zc8orKegW0UsBTt0eU-hPwamfWk7quAtcBIg9fQ;",
        "Upgrade-Insecure-Requests": "1",
        "Pragma": "no-cache",
        "Cache-Control": "no-cache",
        "TE": "Trailers",
    }
    response = getHTML(url,headers)
    soup = BeautifulSoup(response,features="html.parser")
    namesHtml = soup.find_all('div',{'class':'_46pzdl'})
    links = []
    for item in namesHtml:
        links.append(item.find('div',{'class':'_1h3cgic'}).find('a',{'class':'_13ptbeu'}).get('href'))
        try:
            linkToBranches = item.find('span',{"class":"_14quei"}).find('a',{'class':'_13ptbeu'}).get('href')
            if linkToBranches != None:
                links.extend(sarchBranches("https://2gis.ru"+linkToBranches+"?"+coordinates))
        except Exception as e:
            print('None branches')
    return links

def parsingPriview(result:dict):
    data = {}
    try:
        data['id'] = result['id']
    except KeyError as e:
        return None
    try:
        data['name'] = result['name']
    except KeyError as e:
        return None
    try:
        data['address'] = result['address']
    except KeyError as e:
        return None
    try:
        data['address_name'] = result['address_name']
    except KeyError as e:
        return None 
    try:
        data['city'] = result['adm_div'][0]['name']
    except KeyError as e:
        return None
    try:
        data['point'] = result['point']
    except KeyError as e:
        return None
    try:
        data['text'] = result['ads']['text']
    except KeyError as e:
        data['text'] = 'non-data'
    try:
        data['article'] = result['ads']['article']
    except KeyError as e:
        data['article'] = 'non-data'
    try:
        data['rating'] = result['reviews']['general_rating']
    except KeyError as e:
        data['rating'] = 'non-data'
    try:
        data['org_id'] = result['org']['id']
    except KeyError as e:
        return None   
    try:
        data['branch_count'] = result['org']['branch_count']
    except KeyError as e:
        return None 
    return data

def getWorkBook(urlFile: str)->openpyxl.workbook.workbook.Workbook:
    """Load work book .xlsx"""
    try:
        workBook = openpyxl.load_workbook(filename=urlFile)
    except openpyxl.utils.exceptions.InvalidFileException:
        raise SystemExit("Formant file .xls not support. Use formal .xlsx")
    except FileNotFoundError:
        raise SystemExit('File not fount in derectory')
    return workBook

def getWorkSheet(workBook: openpyxl.workbook.workbook.Workbook)->openpyxl.worksheet.worksheet.Worksheet:
    """Load first sheet from book"""
    try:
        sheet = workBook[workBook.sheetnames[0]]
        return sheet
    except TypeError as e:
        sheet = workBook.sheet_by_index(0)
        return e

def appendToFile(urlFile: str, data: list)->dict:
    """Writing new lines ti the end of the file"""
    workBook = getWorkBook(urlFile)
    sheet = getWorkSheet(workBook)
    i = sheet.max_row+1
    for item in data:
        try:
            sheet.cell(row=i, column=1).value = item
            i += 1
        except Exception as e:
            print(e)
    try:
        workBook.save(urlFile)
        workBook.close()
    except Exception as e:
        print(e)
    return {
        'urlFile':urlFile,
        'All rows:':sheet.max_row,
    }

def main():
    i = 0
    subjectsFile = open('data.json')
    subjects = json.loads(subjectsFile.read())
    organizations = []
    for item in subjects['data']:
        try:
            parsingOnePlace(item['name'],item['coordinaes'])
        except Exception as e:
            print(e)
        i+=1
        print(i,'/37')

if __name__ == '__main__':
    main()